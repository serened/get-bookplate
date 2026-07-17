"""Generate the Bookplate home-library-inventory XLSX template.

Palette (Bookplate style guide, exactly four + ink):
burgundy 561A27, coral D75340, green A4A056, cream F2EEE8, ink 1A120A.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

BURGUNDY = "561A27"
CORAL = "D75340"
GREEN = "A4A056"
CREAM = "F2EEE8"
INK = "1A120A"

OUT_DIR = os.path.join(os.path.dirname(__file__), "library-template")
os.makedirs(OUT_DIR, exist_ok=True)

HEADERS = [
    "ISBN", "Title", "Author", "Genre", "Format",
    "Publisher", "Publication year", "Pages", "Location", "Acquired",
    "Status", "Rating", "Date read", "My notes",
]
# Three groups, mirroring the app's own structure:
#   A-H (burgundy)  = the book
#   I-J (coral)     = your copy — where it lives, when it joined the shelves
#   K-N (green)     = your reading — status, rating, date read, notes; on
#                     import these fill the importing user's journal
# No copy-level Notes column (ruling 2026-07-17: the importer has no home for
# it; "My notes" is reading notes and lands in the journal). Extra user-added
# columns are safe — the importer ignores unknown headers. Acquired and Date
# read accept a year ("2019") or a full date ("2019-12-25").
BOOK_COLS = 8
COPY_COLS = 2
WIDTHS = [17, 32, 22, 16, 14, 22, 16, 8, 16, 13, 18, 8, 13, 40]

STARTERS = [
    ["9780141439518", "Pride and Prejudice", "Jane Austen", "Fiction",
     "Softcover", "Penguin Classics", 2002, 480, "Living room", "2015",
     "Read", 5, "2016-01-10", "Reread every few winters."],
    ["9780547928227", "The Hobbit", "J. R. R. Tolkien", "Fantasy",
     "Softcover", "Mariner Books", 2012, 300, "Living room", "2019-12-25",
     "Read", 4, "2020-01-06", "Read aloud over Christmas break."],
    ["9780451524935", "1984", "George Orwell", "Fiction",
     "Softcover", "Signet Classics", 1961, 328, "Office", "2021-03-02",
     "Read", 4, "2021-03-15", ""],
    ["9781400033416", "Beloved", "Toni Morrison", "Fiction",
     "Hardcover", "Vintage", 2004, 324, "Bedroom", "2024-05-14",
     "Currently reading", "", "", "Book club pick for May."],
]

GENRES = ["Fiction", "Nonfiction", "Mystery", "Science fiction", "Fantasy",
          "Romance", "Biography", "History", "Poetry", "Children's"]
LOCATIONS = ["Living room", "Bedroom", "Office", "Hallway", "Storage"]
FORMATS = ["Hardcover", "Softcover"]
STATUSES = ["Read", "Currently reading", "To read", "Did not finish"]

LAST_ROW = 500  # dropdowns and ISBN text format reach this far

wb = Workbook()

header_font = Font(bold=True, color=CREAM, size=11)
reading_font = Font(bold=True, color=INK, size=11)
book_fill = PatternFill("solid", fgColor=BURGUNDY)
copy_fill = PatternFill("solid", fgColor=CORAL)
reading_fill = PatternFill("solid", fgColor=GREEN)
thin_cream = Border(bottom=Side(style="thin", color=CREAM))

# ---- My library ----
ws = wb.active
ws.title = "My library"

for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=i, value=h)
    if i <= BOOK_COLS:
        c.font, c.fill = header_font, book_fill
    elif i <= BOOK_COLS + COPY_COLS:
        c.font, c.fill = header_font, copy_fill
    else:
        c.font, c.fill = reading_font, reading_fill
    c.alignment = Alignment(vertical="center")
    c.border = thin_cream
    ws.column_dimensions[chr(64 + i)].width = WIDTHS[i - 1]
ws.row_dimensions[1].height = 22

for r, row in enumerate(STARTERS, start=2):
    for i, v in enumerate(row, start=1):
        ws.cell(row=r, column=i, value=v)

# ISBNs must stay text so ISBN-10 leading zeros survive
for r in range(2, LAST_ROW + 1):
    ws.cell(row=r, column=1).number_format = "@"

# Acquired and Date read stay text so year-only entries ("2015") survive
# Excel's date coercion
for r in range(2, LAST_ROW + 1):
    ws.cell(row=r, column=10).number_format = "@"
    ws.cell(row=r, column=13).number_format = "@"

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:N{LAST_ROW}"

def dropdown(target_range, list_range):
    dv = DataValidation(type="list", formula1=f"'Lists'!{list_range}",
                        allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv)
    dv.add(target_range)

dropdown(f"D2:D{LAST_ROW}", "$A$2:$A$41")   # Genre
dropdown(f"E2:E{LAST_ROW}", "$C$2:$C$41")   # Format
dropdown(f"I2:I{LAST_ROW}", "$B$2:$B$41")   # Location
dropdown(f"K2:K{LAST_ROW}", "$D$2:$D$41")   # Status

rating_dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5",
                           allow_blank=True, showErrorMessage=False)
ws.add_data_validation(rating_dv)
rating_dv.add(f"L2:L{LAST_ROW}")

# ---- Lists ----
ls = wb.create_sheet("Lists")
for col, (title, values) in enumerate(
        [("Genres", GENRES), ("Locations", LOCATIONS), ("Formats", FORMATS),
         ("Statuses", STATUSES)],
        start=1):
    c = ls.cell(row=1, column=col, value=title)
    c.font = header_font
    c.fill = book_fill
    c.border = thin_cream
    for r, v in enumerate(values, start=2):
        ls.cell(row=r, column=col, value=v)
    ls.column_dimensions[chr(64 + col)].width = 20
ls.row_dimensions[1].height = 22
note = ls.cell(row=1, column=5,
               value="These lists feed the dropdowns on the My library tab. "
                     "Edit them to match your shelves.")
note.font = Font(italic=True, color=INK)
ls.column_dimensions["E"].width = 60

# ---- Start here ----
sh = wb.create_sheet("Start here")
sh.sheet_view.showGridLines = False
sh.column_dimensions["B"].width = 90
title = sh.cell(row=2, column=2, value="Home library inventory")
title.font = Font(name="Georgia", size=18, bold=True, color=BURGUNDY)
steps = [
    "1. Add one row per book on the My library tab. The ISBN is on the back cover or the copyright page.",
    "2. Genre, Format, Location, and Status are dropdowns. Edit the choices on the Lists tab to match your shelves.",
    "3. The green columns are your reading life: status, rating, when you read it, what you thought.",
    "4. Keep the header row as it is. Sort and filter from row 2 down.",
    "5. The example rows show the idea. Replace them with your own books.",
]
for i, s in enumerate(steps, start=4):
    c = sh.cell(row=i, column=2, value=s)
    c.font = Font(size=12, color=INK)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    sh.row_dimensions[i].height = 20
credit = sh.cell(row=10, column=2,
                 value="Made by Bookplate (bookplate.app). If the spreadsheet ever starts to feel "
                       "like homework, import this exact file and your covers and book details "
                       "fill in for you.")
credit.font = Font(size=11, italic=True, color=CORAL)
credit.alignment = Alignment(wrap_text=True, vertical="top")
sh.row_dimensions[10].height = 34

path = os.path.join(OUT_DIR, "home-library-inventory.xlsx")
wb.save(path)
print("wrote", path)
